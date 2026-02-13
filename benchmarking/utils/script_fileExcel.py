import csv

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# FILE_UTENTI = "UTENTI HUMMUS.txt"
FILE_UTENTI = "utenti_h.txt"

FILE_FCQ = "./questionnaire_FCQ.csv"
FILE_JC = "./questionnaire_JC.csv"
FILE_OUTPUT_FCQ = "questionari_FCQ.xlsx"
FILE_OUTPUT_JC = "questionari_JC.xlsx"


def load_users_txt(filename):
    utenti_map = {}
    try:
        with open(filename, encoding="utf-8") as f:
            for riga in f:
                parti = riga.strip().split("\t")
                if len(parti) >= 4:
                    id_str = parti[0]
                    bio_str = parti[3]
                    try:
                        utenti_map[int(id_str)] = bio_str
                    except ValueError:
                        print(f"Warning: Non-numeric ID ignored: {id_str}")
        if not utenti_map:
            print(f"WARNING: No users loaded from {filename}.")
        return utenti_map
    except FileNotFoundError:
        print(f"ERROR: File not found: {filename}")
        return {}


def load_csv_data(filename):
    try:
        with open(filename, encoding="utf-8") as f:
            return list(csv.DictReader(f))
    except FileNotFoundError:
        print(f"ERROR: File not found: {filename}")
        return []


def create_excel_file(output_filename, utenti_list, domande_list, question_type):
    print(f"\nCreating file '{output_filename}'...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = question_type

    # Styles
    grassetto = Font(bold=True)
    stile_intestazione = Alignment(wrap_text=True, vertical="center", horizontal="left")
    stile_testo = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # Base headers
    headers = ["user_id", "biography_text"]

    # Add specific questions
    if question_type == "FCQ":
        for d in domande_list:
            domanda = d.get("questions", "MISSING QUESTION")
            headers.append(domanda)
    elif question_type == "JC":
        for d in domande_list:
            domanda = d.get("questions", "MISSING QUESTION")
            opzioni = d.get("options", "").strip()
            if opzioni:
                domanda = f"{domanda}\n(Options: {opzioni})"
            headers.append(domanda)

    # Writing headers and formatting columns
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = grassetto
        cell.alignment = stile_intestazione

        # column widths
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 10
        elif col_idx == 2:
            ws.column_dimensions[col_letter].width = 35
        else:
            # Standard width for question columns
            ws.column_dimensions[col_letter].width = 35

    ws.row_dimensions[1].height = 100  # height for text wrap

    # Inserting users
    riga = 2
    for utente in utenti_list:
        ws.cell(row=riga, column=1, value=utente["id"])
        cell_bio = ws.cell(row=riga, column=2, value=utente["bio"])
        cell_bio.alignment = stile_testo  # Apply style only to bio
        riga += 1

    # File saving
    try:
        wb.save(output_filename)
        print(f"File '{output_filename}' created successfully!")
        print(f"Inserted {len(utenti_list)} users and {len(headers) - 2} questions.")
    except PermissionError:
        print(f"ERROR: Unable to save '{output_filename}'.")
        print("Check that the file is not open in another program.")
    except Exception as e:
        print(f"Unknown ERROR while saving '{output_filename}': {e}")


# Data loading
print(f"Loading users from {FILE_UTENTI}...")
utenti_map = load_users_txt(FILE_UTENTI)
utenti = [{"id": uid, "bio": bio} for uid, bio in utenti_map.items()]
utenti.sort(key=lambda u: u["id"])
print(f"Found {len(utenti)} users.")

print("Loading questions...")
domande_fcq = load_csv_data(FILE_FCQ)
domande_jc = load_csv_data(FILE_JC)


if not utenti:
    print("No users found. Unable to create files.")
else:
    # Create FCQ file
    if domande_fcq:
        create_excel_file(FILE_OUTPUT_FCQ, utenti, domande_fcq, "FCQ")
    else:
        print(
            f"No FCQ questions found in {FILE_FCQ}. Skipping creation of {FILE_OUTPUT_FCQ}."
        )

    # Create JC file
    if domande_jc:
        create_excel_file(FILE_OUTPUT_JC, utenti, domande_jc, "JC")
    else:
        print(
            f"No JC questions found in {FILE_JC}. Skipping creation of {FILE_OUTPUT_JC}."
        )

print("\nProcessing completed.")
